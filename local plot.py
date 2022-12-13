import io, os
import queue
import time
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from Google import Create_Service
import pandas as pd
import threading as th
import matplotlib.pyplot as plt
import numpy as np
import scipy
from scipy.special import erf
from datetime import datetime
import colorsys
import sigfig
import sympy
from sympy import oo
from scipy.signal import lfilter
from numpy import ma
from openpyxl import load_workbook
from colorama import Fore
from pyampd.ampd import find_peaks, find_peaks_adaptive


CLIENT_SECRET_FILE = 'credentials.json'  # file storing API secret and OAuth2 data
API_NAME = 'drive'
API_VERSION = 'v3'
SCOPES = ['https://www.googleapis.com/auth/drive']
service = Create_Service(CLIENT_SECRET_FILE, API_NAME, API_VERSION, SCOPES)  # Launch connection to Google Drive
folder_id = '1tDl1kuyby0I4bIWasvJW4ve0e_VIBZhF'  # Id of import folder
folder_id2 = '1psz16HbUXYlQ12T9476trbyL33-u6h8l'
keep_going = True  # Variable used to detect keyboard interupt


def key_capture_process():  # Thread used in parallel to specifically check for keyboard interupt and safely abort program, not used due to compatability
    global keep_going
    input()
    keep_going = False


def main():  # Main thread used to launch other parallel tasks
    plot_queue = queue.Queue()  # Method to share data safely from Google Drive thread to Plotting Thread
    lock = th.Semaphore()  # Semaphore to prevent thread clashing
    # th.Thread(target=key_capture_process, args=(), name='key_capture_process', daemon=True).start()   #Thread used in parallel to specifically check for keyboard interupt and safely abort program, caused compatability issues
    th.Thread(target=g_drive_daemon, args=(plot_queue, service, folder_id, lock),
              name='google_drive_daemon').start()  # Thread used in parallel to check for new files in Google Drive
    th.Thread(target=plot_hub, args=(plot_queue, service, folder_id2, lock),
              name='plot_thread').start()  # Thread used in parallel to plot data, convert file types and upload data


def g_drive_daemon(plot_queue, service, folder_id,
                   lock):  # Thread used in parallel to check for new files in Google Drive
    query = f"parents = '{folder_id}' and trashed = false"  # Query to send to Google Drive API
    try:
        with open(os.path.join('./tmp/Archive.txt'),
                  "r") as f:  # Crash mitigation file (saved after every drive search, shows already plotted data files)
            lines = f.read().splitlines()
            f.close()
        store_df = pd.DataFrame(np.zeros(shape=(len(lines), 4)), columns=["kind", "id", "name", "mimeType"])
        store_df['id'] = np.resize(lines, len(store_df))
    except:
        store_df = pd.DataFrame(
            columns=["kind", "id", "name", "mimeType"])  # For when Archive.txt is empty or not yet created
    request_time = 0.5
    while keep_going:  # Will repeat until keyboard interupt, will gracefully exit
        try:
            time.sleep(request_time)  # Time between API requests as to not overload the API
            lock.acquire()  # Semaphore used to only allow one thread to send API requests at once
            response = service.files().list(q=query).execute()  # Send query to API (What files in import folder?)
            files = response.get('files')  # Data from request
            #print(response)
            nextPageToken = response.get('nextPageToken')
            while nextPageToken:  # Used if more than a page of files
                response = service.files().list(q=query, pageToken=nextPageToken).execute()
                #print(response)
                files.extend(response.get('files'))
                nextPageToken = response.get('nextPageToken')
            lock.release()
            pd.set_option('display.max_columns', 100)
            df = pd.DataFrame(
                files)  # Stores file metadata in a Pandas Dataframe for analysis (Basically Python Spreadsheet)
            if not df.equals(store_df):  # Detects if there are new files
                for index, row in df.iterrows():
                    if row.values[1] not in store_df.values and row.values[1] != "application/octet-stream":
                        plot_queue.put((str(row.values[1]), str(row.values[
                                                                    2])))  # 2nd level check to check individual files, if files are different, then the file ID and name are sent for processing with the Plot Thread
                    elif row.values[2] not in store_df.values and row.values != "application/octet-stream":
                        plot_queue.put((str(row.values[2]), str(row.values[
                                                                    3])))


                with open(os.path.join('./tmp/Archive.txt'),
                          "w") as f:  # Save already requested plots to prevent duplicate plots at a crash
                    for index, row in df.iterrows():
                        if row.values[1] != "application/octet-stream":
                            f.write(row.values[1] + '\n')  # Only ID is needed to detect new files
                        elif row.values[2] != "application/octet-stream":
                            f.write(row.values[2] + '\n')
                    f.close()
            store_df = df
            request_time = 0.5  # Predict that the API now works again and can take requests at the usual rate
        except:  # System to increase request time if the Google API starts to fail
            if request_time < 5:
                request_time += 0.1


def plot_hub(plot_queue, service, folder_id2,
             lock):  # Thread used in parallel to plot data, convert file types and upload data
    plt.tick_params(bottom=True, top=True, left=True, right=True, direction='in')  # Matplotlib graph styling
    plt.tick_params(labelbottom=True, labeltop=False, labelleft=True, labelright=False, direction='in')
    plt.rcParams['text.usetex'] = True
    plt.rcParams['font.size'] = '16'
    fig = plt.figure(figsize=(21, 9))
    ax = fig.add_subplot()
    ax.tick_params(bottom=True, top=True, left=True, right=True, direction='in')
    ax.tick_params(labelbottom=True, labeltop=False, labelleft=True, labelright=False, direction='in')
    for label in (ax.get_xticklabels() + ax.get_yticklabels()):
        label.set_fontsize(16)
    while keep_going:  # Will repeat until keyboard interrupt, will gracefully exit (Now redundant)
        output_intergration = []
        output_efficiency = []
        output_radioiso = []
        output_peak = []
        output_fwhm = []
        id, name = plot_queue.get()  # Retrieve data from Google Drive Thread
        if name[-4:] == ".TKA":  # Prevents wrong files being plotted
            lock.acquire()  # Only one thread can do API requests
            request = service.files().get_media(fileId=id)  # Prepping Google Drive API to download files
            fh = io.BytesIO()  # Prepping local computer to take downloads
            downloader = MediaIoBaseDownload(fh, request)  # Creates download environment
            name = name[:-4]
            done = False
            while not done:  # Download the data
                status, done = downloader.next_chunk()
            fh.seek(0)
            lock.release()  # No longer using Google API, allow other threads to use it
            with open(os.path.join('./tmp/local_data.csv'),
                      "wb") as f:  # Save the .TKA file in an easier to work with format (locally)
                f.write(fh.read())
                f.close()
            master_data = pd.read_csv('./tmp/local_data.csv')  # Create internal spreadsheet of data
            y = master_data.dropna().reset_index(drop=True)
            y = y.values.ravel()
            times = y[:2]
            n = 15  # the larger n is, the smoother curve will be
            b = [1.0 / n] * n
            a = 1
            y = y[2:]
            x = []
            for value, _ in enumerate(y):
                x.append(value + 1)
            x = np.array(x)
            x_store = x
            #x = mca_to_kev(x)
            ax.plot(x, y, '--', marker='+')
            peaks, _ = scipy.signal.find_peaks(y, height=100, prominence= 100)
            colour = (1, 0, 0)
            col = np.array([0, 65536, 65536], dtype=np.uint16)
            change = np.array([65536 / len(peaks), 0, 0], dtype=np.uint8)
            print(len(peaks))
            for peak in peaks:
                #ax.vlines(peak, ymin=min(y) * 0.95, ymax=max(y) * 1.05, color=colour,
                #          label=peak)

                lower = 0
                higher = 0
                repeat = True
                repeat_l = True
                repeat_r = True
                fail = False
                index = 0
                p0 = (1, x_store[peak], 2, 0.1, y[peak]/2)
                old_cv = False
                old_params = False
                old_mean_l = 0.0
                old_mean_r = 0.0
                if peak == 0 or len(y) -5 < peak:
                    repeat = False
                    fail = True
                else:
                    if y[peak] >= y[peak - 1] and y[peak] >= y[peak +1] and y[peak] > y[peak-2] and y[peak] > y[peak + 2] :
                        higher = 2
                        lower = 2
                    else:
                        repeat = False
                        fail = True
                found = False
                while repeat:

                    if len(y[peak - index - 3: peak-index]) < 3 or len(y[peak + index: peak + index + 3]) < 3 or len(y[peak - index - 6: peak-index -3]) < 3 or len(y[peak + index + 3: peak + index + 6]) < 3:
                        repeat = False
                        break
                    if not repeat_l and not repeat_r:
                        repeat = False
                        break
                    if y[peak-lower-1] < y[peak-lower]:
                        lower += 1
                    elif y[peak-lower-2] < y[peak-lower]:
                        lower += 2
                    elif y[peak-lower-3] < y[peak-lower]:
                        lower += 3
                    else:
                        repeat_l = False
                    if y[peak+higher+1] < y[peak+higher]:
                        higher += 1
                    elif y[peak+higher+2] < y[peak+higher]:
                        higher += 2
                    elif y[peak+higher+3] < y[peak+higher]:
                        higher += 3
                    else:
                        repeat_r = False
                    x_gauss = x_store[peak - lower:peak + higher]
                    x_lead = x_store[peak-lower:peak + 3]
                    y_gauss = y[peak - lower:peak + higher]
                    y_lead = y[peak - lower:peak + 3]
                    #print(len(y_gauss))
                    if len(y_gauss) > 10 and lower > 4 and higher > 4:
                        #print(len(y_gauss))
                        #print(len(x_gauss))
                        fitting_funcs = [leading_edge, gaussian, split_gaussian, skewed_gaussian]
                        plotting_funcs = [gauss, gauss, split_gauss, skewed_gauss]
                        method = 0
                        force_cont = False
                        num_sigma = 1
                        min_num_sigma = False
                        rtest = 0.9
                        params, cv, good = leading_edge(x_lead, y_lead, x[peak], rtest)
                        if good:
                            num_sigma = abs(x[peak]-params[1])
                            min_max = (min(x_lead), max(x_lead), min(y_lead), max(y_lead))
                            if num_sigma <= 1:
                                min_num_sigma = [method, num_sigma, params, cv, min_max]
                                method += 1
                            else:
                                min_num_sigma = [method, num_sigma, params, cv, min_max]
                        else:
                            force_cont = True
                        method = 1
                        while num_sigma > 1 or force_cont:
                            if method == 4:
                                break
                            force_cont = False
                            params, cv, good = fitting_funcs[method](x_gauss, y_gauss, x[peak], rtest)
                            if good:
                                num_sigma = abs(x[peak] - params[1])
                                min_max = (min(x_gauss), max(x_gauss), min(y_gauss), max(y_gauss))
                                if num_sigma <= np.sqrt(np.diag(cv))[1]:
                                    min_num_sigma = [method, num_sigma, params, cv, min_max]
                                    method += 1
                                elif not isinstance(min_num_sigma, bool):
                                    if num_sigma < min_num_sigma[1]:
                                        min_num_sigma = [method, num_sigma, params, cv, min_max]
                                        method += 1
                                    else:
                                        method += 1
                                elif isinstance(min_num_sigma, bool) and num_sigma <=3:
                                    min_num_sigma = [method, num_sigma, params, cv, min_max]
                                    method += 1
                                else:
                                    method += 1
                            else:
                                method += 1
                        if not isinstance(min_num_sigma, bool):
                            if min_num_sigma[1] <=3:
                                store_min_num_sigma = min_num_sigma
                                found = True
                if found:
                    if store_min_num_sigma[0] <=1:
                            ax.vlines(store_min_num_sigma[2][1], ymin=min(y) * 0.95, ymax=max(y) * 1.05, color=colour, label="(" + str(sigfig.round(store_min_num_sigma[2][1],  np.sqrt(np.diag(store_min_num_sigma[3]))[1])) + ") Counts")
                            x_new = np.linspace(store_min_num_sigma[4][0], store_min_num_sigma[4][1])
                            func = plotting_funcs[store_min_num_sigma[0]](x_new, store_min_num_sigma[2][0], store_min_num_sigma[2][1], store_min_num_sigma[2][2], store_min_num_sigma[2][3])
                            plt.plot(x_new, func, color="k")
                            area = store_min_num_sigma[2][0] * store_min_num_sigma[2][2]/ (np.sqrt(2 * np.pi))
                        #except:
                        #    print("Fail",store_min_num_sigma[2][1],np.diag(store_min_num_sigma[3][1]))
                    elif store_min_num_sigma[0] == 2:
                        try:
                            ax.vlines(store_min_num_sigma[2][1], ymin=min(y) * 0.95, ymax=max(y) * 1.05, color=colour,
                                      label="(" + str(sigfig.round(store_min_num_sigma[2][1],
                                                                   np.sqrt(np.diag(store_min_num_sigma[3]))[1])) + ") Counts")
                            x_new = np.linspace(store_min_num_sigma[4][0], store_min_num_sigma[4][1])
                            func = plotting_funcs[store_min_num_sigma[0]](x_new, store_min_num_sigma[2][0],
                                                                          store_min_num_sigma[2][1],
                                                                          store_min_num_sigma[2][2],
                                                                          store_min_num_sigma[2][3], store_min_num_sigma[2][4], store_min_num_sigma[2][5])
                            plt.plot(x_new, func, color="k")
                            area = store_min_num_sigma[2][0] * store_min_num_sigma[2][3] / (np.sqrt(2 * np.pi))*0.5 + store_min_num_sigma[2][2] * store_min_num_sigma[2][4] / (np.sqrt(2 * np.pi))*0.5
                        except:
                            print("Fail",store_min_num_sigma[2][1],np.diag(store_min_num_sigma[3][1]))

                    elif store_min_num_sigma[0] == 3:
                        ax.vlines(store_min_num_sigma[2][1], ymin=min(y) * 0.95, ymax=max(y) * 1.05, color=colour, label="(" + str(sigfig.round(store_min_num_sigma[2][1],  np.sqrt(np.diag(store_min_num_sigma[3]))[1])) + ") Counts")
                        x_new = np.linspace(store_min_num_sigma[4][0], store_min_num_sigma[4][1])
                        func = plotting_funcs[store_min_num_sigma[0]](x_new, store_min_num_sigma[2][0], store_min_num_sigma[2][1], store_min_num_sigma[2][2], store_min_num_sigma[2][3], store_min_num_sigma[2][4])
                        plt.plot(x_new, func, color="k")
                    try:
                        print(area)
                    except:
                        print("skewed")
                    #N_peak = gauss_integrate(x_coni, params_con, np.mean([old_y[0], old_y[-1]]), errors_con)
                    #output_intergration.append(N_peak)
                    #efficiency = efficiency_curve(x_con)
                    #output_peak.append([x_con, x_conu])
                    #output_efficiency.append(efficiency)
                    #output_fwhm.append([x_consig * 2 * np.sqrt(2 * np.log(2)), x_consigu * 2 * np.sqrt(2 * np.log(2))])
                    #radioiso = isotope_detector(x_con, x_conu)
                    #output_radioiso.append(radioiso)
                    #plt.text(old_params[1], max(y) * 0.9,
                             #"" + str(sigfig.round(old_params[1], np.sqrt(np.diag(cv))[1])) + "", color=colour,
                             #rotation=90)
                col += change
                colour = colorsys.hsv_to_rgb(float(col[0]) / 65536, 1, 1)

            output = []
            #output = [name]
            #for value, item in enumerate(output_peak):
            #    output.append("Peak Energy:")
            #    output.append("(" + str(sigfig.round(output_peak[value][0],  output_peak[value][1])) + ") KeV")
            #    output.append("FWHM:")
            #    output.append("(" + str(sigfig.round(output_fwhm[value][0],  output_fwhm[value][1])) + ") KeV")
            #    output.append("Area of photopeak:")
            #    try:
            #        output.append(str(sigfig.round(output_intergration[value][0],  output_intergration[value][1])))
            #    except:
            #        output.append("Error")
            #    output.append("Efficiency")
            #    output.append(str(output_efficiency[value]))
            #    output.append("Potential radioisotopes:")
            #    for isotope in output_radioiso[value]:
            #        output.append(isotope)
            #    output.append("-----------------------------------------------------------------------------------------\n")
            with open(r'./tmp/' + name + "_plotting_variables.txt" , 'w') as fp:
                for item in output:
                    # write each item on a new line
                    fp.write("%s\n" % item)

            plt.ylabel(r"Counts", fontsize=20)
            plt.xlabel("Energy/KeV", fontsize=20)
            plotname = './tmp/' + name + '.png'  # Creating output plot file name
            plt.legend(fontsize=8)
            ax.set_yscale("log")
            plt.savefig(plotname, dpi=300)
            #exit()  # Save plot
            ax.clear()
            excelname = './tmp/' + name + '.xlsx'  # Creating output excel data file name
            master_data.to_excel(excelname)  # Save excel
            filenames = [name + ".xlsx", name + ".png", name + "_plotting_variables.txt"]
            mime_types = ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                          'image/png', 'text/plain']  # Define filetypes for Google Drive API
            lock.acquire()  # Protect Google API against simultaneous usage
            for file, mime in zip(filenames, mime_types):  # Upload data to Google Drive in output folder
                file_metadata = {
                    'name': file,
                    'parents': [folder_id2]
                }
                media = MediaFileUpload('./tmp/{0}'.format(file), mimetype=mime)
                service.files().create(
                    body=file_metadata,
                    media_body=media,
                    fields='id'
                ).execute()
            now = datetime.now()
            finish_time = now.strftime("%A, %d, %m, %Y\n%H:%M:%S")
            print(name + " complete" + "\n" + finish_time + "\n")
            lock.release()  # Allow other threads to access the API


def leading_edge(x, y, peak, r_check):
    amp = peak - min(y)
    if amp < 0:
        amp = 1
    p0 = (amp, peak, 5, min(y))
    params, cv = scipy.optimize.curve_fit(gauss, x, y, p0, maxfev=1000000, bounds=((0,0,0,0),(np.inf, np.inf, np.inf, np.inf)))
    a, mean, sigma, b = params
    squaredDiffs = np.square(y - gauss(x, a, mean, sigma, b))
    squaredDiffsFromMean = np.square(y - np.mean(y))
    rSquared = 1 - np.sum(squaredDiffs) / np.sum(squaredDiffsFromMean)
    if rSquared > r_check:
        return params, cv, True
    else:
        return params, cv, False


def gaussian(x, y, peak, r_check):
    amp = peak-min(y)
    if amp <0:
        amp = 1
    p0 = (amp, peak, 5, min(y))
    params, cv = scipy.optimize.curve_fit(gauss, x, y, p0, maxfev=1000000, bounds=((0,0,0,0),(np.inf, np.inf, np.inf, np.inf)))
    a, mean, sigma, b = params
    squaredDiffs = np.square(y - gauss(x, a, mean, sigma, b))
    squaredDiffsFromMean = np.square(y - np.mean(y))
    rSquared = 1 - np.sum(squaredDiffs) / np.sum(squaredDiffsFromMean)
    if rSquared > r_check:
        return params, cv, True
    else:
        return params, cv, False


def split_gaussian(x, y, peak, r_check):
    amp = peak - min(y)
    if amp <0:
        amp = 1
    p0 = (amp, 1, peak, 5, 5, min(y))
    params, cv = scipy.optimize.curve_fit(split_gauss, x, y, p0, maxfev=1000000, bounds=((0,0,0,0,0,0),(np.inf, np.inf, np.inf, np.inf,np.inf,np.inf)))
    A1, mean, A2, sigma1, sigma2, b= params
    squaredDiffs = np.square(y - split_gauss(x, A1, mean, A2, sigma1, sigma2, b))
    squaredDiffsFromMean = np.square(y - np.mean(y))
    rSquared = 1 - np.sum(squaredDiffs) / np.sum(squaredDiffsFromMean)
    if rSquared > r_check:
        return params, cv, True
    else:
        return params, cv, False


def skewed_gaussian(x, y, peak, r_check):
    amp = peak - min(y)
    if amp <0:
        amp = 1
    p0 = (amp, peak, 5, 1, min(y))
    params, cv = scipy.optimize.curve_fit(skewed_gauss, x, y, p0, maxfev=1000000, bounds=((0,0,0,-np.inf,0),(np.inf, np.inf, np.inf, np.inf, np.inf)))
    A, mu, sigma, gamma, b = params
    squaredDiffs = np.square(y - skewed_gauss(x, A, mu, sigma, gamma, b))
    squaredDiffsFromMean = np.square(y - np.mean(y))
    rSquared = 1 - np.sum(squaredDiffs) / np.sum(squaredDiffsFromMean)
    if rSquared > r_check:
        return params, cv, True
    else:
        return params, cv, False


def split_skewed_gaussian(x, y, peak, r_check): #not in use, prone to complications
    p0 = (1, peak, 5, 500)
    params, cv = scipy.optimize.curve_fit(gauss, x, y, p0, maxfev=1000000, bounds=((0,0,0,0),(np.inf, np.inf, np.inf, np.inf)))
    a, mean, sigma, b = params
    squaredDiffs = np.square(y - gauss(x, a, mean, sigma, b))
    squaredDiffsFromMean = np.square(y - np.mean(y))
    rSquared = 1 - np.sum(squaredDiffs) / np.sum(squaredDiffsFromMean)
    if rSquared > r_check:
        return params, cv, True
    else:
        return params, cv, False


def gauss(x, a, mean, sigma, b):
    return b + a * np.exp(-(x - mean) ** 2 / (2 * sigma ** 2))


def background(x, b, n, c):
    return b * x**n +c

def split_gauss(x, A1, mean, A2, sigma1, sigma2, b1):
    b2 = A1-A2 +b1
    low_x = x[x < mean]
    high_x = x[x >= mean]
    low_gauss = b1 + A1 * np.exp(-(low_x - mean) ** 2 / (2 * sigma1 ** 2))
    high_gauss = b2 + A2 * np.exp(-(high_x - mean) ** 2 / (2 * sigma2 ** 2))
    return np.concatenate((low_gauss, high_gauss))


def split_skewed_gauss(x, A1, A2, mu, sigma, b1, gamma):
    denom = np.sqrt(np.pi) * sigma
    low_x = x[x < mu]
    high_x = x[x >= mu]
    b2 = (A1 / denom - A2 / denom + b1) / high_x
    low_gauss = b1 * low_x +A1/denom * np.exp((-(low_x-mu)**2)/(2 * sigma ** 2)) * (1 + erf((gamma * (low_x-mu))/(sigma * np.sqrt(2)
    )))
    high_gauss = b2 * high_x + A2 / denom * np.exp((-(high_x - mu) ** 2) / (2 * sigma ** 2)) * (1 + erf((gamma * (high_x - mu)) / (sigma * np.sqrt(2))))
    #print(np.concatenate((low_gauss, high_gauss)))
    return np.concatenate((low_gauss, high_gauss))


def skewed_gauss(x, A, mu, sigma,  gamma, b):
    denom = np.sqrt(np.pi) * sigma
    return b + A / denom * np.exp((-(x - mu) ** 2) / (2 * sigma ** 2)) * (1 + erf((gamma * (x - mu)) / (sigma * np.sqrt(2))))



def mca_to_kev(x):
    return 0.276 * x + 0.5 # Replace with found values


def efficiency_curve(x):
    return (2 * 8.66 / 100) / ((x / 103) ** 1.1 + (103 / x) ** 3.6)


def gauss_integrate(x, params, height, errors):
    #print(max(x) - min(x))
    #print(height)
    local_x = sympy.symbols('x')
    a, mean, sigma = params
    local_gauss = a * sympy.exp(-(local_x - mean) ** 2 / (2 * sigma ** 2))
    area = sympy.integrate(local_gauss, (local_x, min(x), max(x))) #- height * (max(x) - min(x))
    print(area)
    return (area, area * np.sqrt((errors[0] / a) ** 2 + (errors[1] / mean) ** 2 + (errors[2] / sigma) ** 2))


#def radioactivity(N_peak, times, energy, uncertainty):
#    efficiency = efficiency_curve(energy)
#    #isotope_detector(energy, uncertainty)
#    return


def isotope_detector(energy, uncertainty):
    master_data = pd.ExcelFile('./Data/Nuclide_data_store.xlsx')
    global sheets
    old_energy_levels = []
    old_intensities = []
    output = []
    nearest = True
    for sheet in sheets:
        df = pd.read_excel(master_data, sheet)
        try:
            max(df['Energy (keV)']) < energy
            min(df['Energy (keV)']) > energy
        except:
            continue
        if len(df['Energy (keV)']) >= 2 and max(df['Energy (keV)']) > energy > min(df['Energy (keV)']):
            energy_levels = find_neighbours(energy, df, colname='Energy (keV)')
            intensities = [df['Intensity (%)'][energy_levels[0]], df['Intensity (%)'][energy_levels[1]]]
        elif len(df['Energy (keV)']) == 1:
            energy_levels = [0, 0]
            intensities = [df['Intensity (%)'][energy_levels[0]], 0]
        elif max(df['Energy (keV)']) < energy and len(df['Energy (keV)']) > 0:
            energy_levels = [df['Energy (keV)'].idxmax(), df['Energy (keV)'].idxmax()]
            intensities = [df['Intensity (%)'][energy_levels[0]], 0]
        elif min(df['Energy (keV)']) > energy and len(df['Energy (keV)']) > 0:
            energy_levels = [df['Energy (keV)'].idxmin(), df['Energy (keV)'].idxmin()]
            intensities = [0, df['Intensity (%)'][energy_levels[1]]]
        else:
            continue
        if nearest == True and len(old_energy_levels) > 1:
            if abs(energy - df['Energy (keV)'][energy_levels[0]]) < abs(energy - dfo['Energy (keV)'][old_energy_levels[0]]) and abs(energy - df['Energy (keV)'][energy_levels[0]]) < abs(energy - dfo['Energy (keV)'][old_energy_levels[1]]) or abs(energy - df['Energy (keV)'][energy_levels[0]]) < abs(energy - dfo['Energy (keV)'][old_energy_levels[0]]) and abs(energy - df['Energy (keV)'][energy_levels[0]]) < abs(energy - dfo['Energy (keV)'][old_energy_levels[1]]):
                old_energy_levels = energy_levels + [sheet]
                old_intensities = intensities
                dfo = df
        elif nearest == True:
            old_energy_levels = energy_levels + [sheet]
            old_intensities = intensities
            dfo = df
        if np.isclose(df['Energy (keV)'][energy_levels[0]], energy, rtol=0.01) or np.isclose(df['Energy (keV)'][energy_levels[1]], energy, rtol=0.01):
            nearest = False
            if len(old_energy_levels) == 0:
                old_energy_levels = energy_levels + [sheet]
                output.append(sheet.strip('.l'))
                old_intensities = intensities
                dfo = df
            else:
                if max(intensities) > max(old_intensities):
                    old_energy_levels = energy_levels + [sheet]
                    output.append(sheet.strip('.l'))
                    old_intensities = intensities
                    dfo = df
    return output










def find_neighbours(value, df, colname):
    exactmatch = df[df[colname] == value]
    if not exactmatch.empty:
        return exactmatch.index
    else:
        lowerneighbour_ind = df[df[colname] < value][colname].idxmax()
        upperneighbour_ind = df[df[colname] > value][colname].idxmin()
        return [lowerneighbour_ind, upperneighbour_ind]


def get_sheetnames_xlsx():
    wb = load_workbook('./Data/Nuclide_data_store.xlsx', read_only=True, keep_links=False)
    return wb.sheetnames

sheets = get_sheetnames_xlsx()
main()